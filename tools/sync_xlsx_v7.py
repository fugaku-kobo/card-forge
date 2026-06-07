# -*- coding: utf-8 -*-
"""手塚カードゲーム v7 — xlsx を docs/rule_v7.md の最新内容に同期する。

対象: data/ 内の最新版 手塚カードゲーム_v7_vX.Y.xlsx
- ページカード/急展開カード/主人公シートの文面を最新化
- ページ/急展開シートのヘッダー「名称」→「カード名」(index.html 取り込み用)
- かんたん/詳細ルールの勝利条件 6→4、リーダー能力、補足ルール追記
再実行可(冪等)。実行時に自動でタイムスタンプ付きバックアップを作成。
実行: py -3 tools/sync_xlsx_v7.py
"""
import openpyxl, os, io, sys
from _common import latest_xlsx, backup

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

# --- ページカード(rule_v7.md N章)効果の最新文面 ---
PAGE_FIX = {
    'PG-001': '戦闘では攻撃力の低い方が勝つ(同値は相殺のまま)',
    'PG-002': '全キャラの攻撃力を「300」にする。その後に適用されるバフ・デバフのみ加減算する',
    'PG-005': '戦闘は各プレイヤーが公開した自分のデッキトップの攻撃力で決める(公開カードは没案へ)',
    'PG-007': '各プレイヤーは自分の没案ゾーンのカード1枚を選び手札に加える',
    'PG-009': '手番プレイヤーは自分のデッキトップ3枚を確認し、好きな順に並べ替えてよい',
    'PG-013': '自分の場のキャラは攻撃力 −200(最低100)',
    'PG-014': '相手の場のキャラは攻撃力 −200(最低100)',
    'PG-030': 'ブロックされずに原稿ゾーンへ得点したプレイヤーは追加で1ドロー(同名重複で没案へ送った場合は対象外)',
}
# --- 急展開カード(O章) ---
KT_EFFECT_FIX = {
    'KT-001': 'このターン、相手の攻撃をすべて無効にする',
    'KT-004': '相手が攻撃した時に発動。攻撃しているキャラを没案ゾーンへ送る(攻撃も止まる)',
    'KT-008': '相手が伏せている急展開カード1枚を、発動される前に没案ゾーンへ送る',
}
KT_NAME_FIX = {'KT-007': '生存フラグ'}
# --- 主人公(K章)両面共通効果 ---
HERO_FIX = {
    'BJ-H001': '「執刀」1ターンに1度、キャラ1体を「執刀」する。自分のキャラなら→このターン戦闘で没にならない+攻撃力+200(救命)。相手のキャラなら→このターン、ブロックできない。',
    'HT-H003': '「輪廻」1ターンに1度、メインフェイズの最初に、没案ゾーンの自分のキャラ1体を場に出す(召喚酔いあり・空き枠が必要)。そのキャラは攻撃力+200(永続)。',
    'RK-H004': '「ふたつの心」サファイアのプレイヤーは「心」を持つ(初期は【王子の心】)。自分のターン開始時に【王子の心】と【お姫さまの心】のどちらかを選ぶ。【王子の心】=自分のキャラは攻撃する時、攻撃力+100。【お姫さまの心】=自分のキャラはブロックする時、攻撃力+200。',
}
# --- キャラクターアイディア: ID -> 技1効果(固有能力)の最新文面 ---
CHAR_FIX = {
    'RK-C001': '自分の心が【王子】なら攻撃力+100、【姫】なら戦闘で没にならない',
    'RK-C005': '自分の心が【王子】なら攻撃力+100、【姫】なら戦闘で没にならない',
    'RK-C017': '自分の心が【王子】の間、攻撃力+200',
    'RK-C021': '自分の心が【王子】の間、攻撃力+200',
}
# --- 詳細ルール I章リーダー行(部分一致で行特定→全文置換) ---
DETAIL_LEADER_FIX = {
    '「執刀」': '・ブラック・ジャック「執刀」:1ターンに1度、キャラ1体を執刀。自分のキャラ＝このターン戦闘で没にならない+攻撃力+200(救命)／相手のキャラ＝このターンブロックできない。(コントロール型)',
    '「輪廻」': '・ナギ(火の鳥)「輪廻」:1ターンに1度、メインフェイズの最初に没案ゾーンの自分のキャラ1体を場に出す(召喚酔いあり・空き枠が必要)。そのキャラは攻撃力+200(永続)。(蘇生型)',
    '「ふたつの心」': '・サファイア(リボンの騎士)「ふたつの心」:「心」を持ち、自分のターン開始時に【王子の心】と【お姫さまの心】のどちらかを選ぶ。【王子】＝自分のキャラは攻撃時に攻撃力+100／【お姫さま】＝自分のキャラはブロック時に攻撃力+200。(変身・適応型)',
}
# --- 詳細ルール末尾に追記する補足ルール ---
DETAIL_APPEND = [
    '',
    '【M. 補足ルール(2026-05-22 追加)】',
    '・進化スタックの移動:進化したキャラ(Lv1+Lv2が重なった状態)が没案・手札・原稿ゾーンへ移動する時、重なったカードはすべて一緒に移動する。',
    '・急展開カードの発動タイミング:「〜した時」と書かれたカード(KT-001/KT-004)はその条件を満たした時のみ発動。それ以外は任意のタイミングで発動でき、自分・相手どちらのターンにも割り込める。',
    '・戦闘判定を変えるページカード(PG-001〜005)の共通裁定:',
    '　- 盾持ち・守り手は代替判定でも有効(攻撃力・ダイスに関係なく攻撃を完全に止める)。ただし貫通を持つアタッカーには無効。',
    '　- 代替判定でアタッカーが勝った場合、通常どおり突破して原稿ゾーンへ(得点)。',
    '　- ダイスが同値(PG-003/004)になったら振り直す。',
    '　- PG-005で公開したカードが急展開カード(攻撃力なし)の場合、攻撃力0として扱う。',
    '・キーワード「貫通」(2026-06-07追加):攻撃時、相手の盾持ち・守り手の「完全に止める」防御を無視し、通常の攻撃力比べに持ち込む(勝てば突破して得点)。代替判定下でも盾無視は有効。',
    '・攻撃力は3桁表記(2026-06-07に全数値×100)。',
]

changes = []


def col_of(ws, header):
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=c).value
        if v is not None and str(v).strip() == header:
            return c
    return None


def sync_card_sheet(ws, effect_fix, name_fix=None):
    nc = col_of(ws, '名称')
    if nc:
        ws.cell(row=1, column=nc).value = 'カード名'
        changes.append('%s: ヘッダー「名称」→「カード名」' % ws.title)
    idc = col_of(ws, 'ID')
    namec = col_of(ws, 'カード名')
    effc = col_of(ws, '効果')
    del_rows = []
    for r in range(2, ws.max_row + 1):
        a = ws.cell(row=r, column=idc).value
        cid = str(a).strip() if a is not None else ''
        if cid.startswith('※'):
            del_rows.append(r)
            continue
        if cid in effect_fix:
            new = effect_fix[cid]
            if ws.cell(row=r, column=effc).value != new:
                ws.cell(row=r, column=effc).value = new
                changes.append('%s %s 効果を更新' % (ws.title, cid))
        if name_fix and cid in name_fix:
            new = name_fix[cid]
            cur = ws.cell(row=r, column=namec).value
            if cur != new:
                ws.cell(row=r, column=namec).value = new
                changes.append('%s %s 名称「%s」→「%s」' % (ws.title, cid, cur, new))
    for r in reversed(del_rows):
        ws.delete_rows(r, 1)
        changes.append('%s: 注記行(※)を削除' % ws.title)


def sync_hero(ws):
    idc = col_of(ws, 'ID')
    effc = col_of(ws, '両面共通効果')
    for r in range(2, ws.max_row + 1):
        a = ws.cell(row=r, column=idc).value
        cid = str(a).strip() if a is not None else ''
        if cid in HERO_FIX:
            new = HERO_FIX[cid]
            if ws.cell(row=r, column=effc).value != new:
                ws.cell(row=r, column=effc).value = new
                changes.append('主人公 %s 効果を更新' % cid)


def sync_char_sheet(ws):
    idc = col_of(ws, 'ID')
    effc = col_of(ws, '技1効果')
    for r in range(2, ws.max_row + 1):
        a = ws.cell(row=r, column=idc).value
        cid = str(a).strip() if a is not None else ''
        if cid in CHAR_FIX:
            new = CHAR_FIX[cid]
            if ws.cell(row=r, column=effc).value != new:
                ws.cell(row=r, column=effc).value = new
                changes.append('キャラ %s 技1効果を更新' % cid)


def sync_rule_sheet(ws, is_detail):
    for r in range(1, ws.max_row + 1):
        cell = ws.cell(row=r, column=1)
        if cell.value is None:
            continue
        s = str(cell.value)
        # 勝利条件 6→4(「勝」または「原稿」を含む行のみ。デッキ構築の同名6枚は除外)
        if ('勝' in s or '原稿' in s) and ('6枚' in s or '6種' in s):
            ns = s.replace('6枚', '4枚').replace('6種', '4種')
            if ns != s:
                cell.value = ns
                s = ns
                changes.append('%s: 勝利条件 6→4 (R%d)' % (ws.title, r))
        # 詳細ルール: I章リーダー行を全文置換
        if is_detail:
            for key, newline in DETAIL_LEADER_FIX.items():
                if key in s and s != newline:
                    cell.value = newline
                    changes.append('%s: リーダー行 %s を更新' % (ws.title, key))
                    break


def main():
    xlsx = latest_xlsx()
    print('対象:', xlsx)
    print('バックアップ:', backup(xlsx))

    wb = openpyxl.load_workbook(xlsx)
    sync_card_sheet(wb['ページカード'], PAGE_FIX)
    sync_card_sheet(wb['急展開カード'], KT_EFFECT_FIX, KT_NAME_FIX)
    sync_hero(wb['主人公'])
    sync_char_sheet(wb['キャラクターアイディア'])
    sync_rule_sheet(wb['かんたんルール'], False)
    ws_detail = wb['詳細ルール']
    sync_rule_sheet(ws_detail, True)
    # 詳細ルール末尾に補足ルールを追記(冪等: 既にあれば追記しない)
    existing = [str(ws_detail.cell(row=r, column=1).value or '')
                for r in range(1, ws_detail.max_row + 1)]
    if not any('【M. 補足ルール' in x for x in existing):
        base = ws_detail.max_row
        for i, line in enumerate(DETAIL_APPEND, start=1):
            ws_detail.cell(row=base + i, column=1).value = line
        changes.append('詳細ルール: 補足ルール(M章)を追記')

    wb.save(xlsx)
    print('--- 変更点 %d 件 ---' % len(changes))
    for c in changes:
        print('  ', c)
    if not changes:
        print('  (変更なし: 既に最新)')
    print('保存完了:', xlsx)


if __name__ == '__main__':
    main()
