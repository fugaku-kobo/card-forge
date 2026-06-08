# -*- coding: utf-8 -*-
"""手塚カードゲーム — 最新 xlsx から card-forge 同梱シード data/seed.js を生成。

アプリ(index.html)のカードモデルへ変換し、window.CARD_FORGE_SEED に
{ version, generatedFrom, cards, types, decks, cardIdCounter } を代入する JS を出力。
index.html は <script src="data/seed.js"> で読み込み、初回(localStorageが空)に自動シードする。

実行: py -3 tools/build_seed.py
"""
import openpyxl, json, io, sys, os, re
from _common import latest_xlsx

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

# 名前→絵文字(index.html の V7_EMOJI 相当・アイコン絵文字列が空のときの補完)
EMOJI = {
    'ブラック・ジャック': '🩺', 'ピノコ': '🧒', '本間丈太郎': '👨‍⚕️', 'ドクター・キリコ': '💀',
    'スパイダー': '🕷️', 'ヒョウタンツギ': '🐽', '鉄腕アトム': '🤖', 'ウラン': '👧', 'コバルト': '🦾',
    'お茶の水博士': '👨‍🔬', '火の鳥': '🔥', 'ナギ': '🧑', 'ナギC': '🧑', '猿田彦': '🗿', '天弓彦': '🏹',
    'サファイア': '👑', 'チンク': '👼', 'フランツ王子': '🤴', 'ヘケート': '🧙‍♀️', 'ロック': '🎯', '写楽': '🎭',
}
ZEN, HAN = '０１２３４５６７８９', '0123456789'


def nz(v):
    return '' if v is None else str(v).strip()


def numstr(v):
    return nz(v).translate(str.maketrans(ZEN, HAN))


def col_of(ws, header):
    for c in range(1, ws.max_column + 1):
        if nz(ws.cell(row=1, column=c).value) == header:
            return c
    return None


def main():
    xlsx = latest_xlsx()
    base = os.path.basename(xlsx)
    ver = 'v' + re.search(r'_v(\d+\.\d+)\.xlsx', base).group(1)
    print('対象:', base, '版:', ver)
    wb = openpyxl.load_workbook(xlsx, data_only=True)

    cards = []
    seq = [0]
    def new_id():
        seq[0] += 1
        return 'c' + str(seq[0])

    def icon_for(name, raw):
        raw = nz(raw)
        return raw if raw else EMOJI.get(name, '')

    chars_by_work = {}   # work -> [card,...]
    hero_by_work = {}    # work -> card id
    kt_id = {}           # KT-002/003/005 -> id

    # --- キャラクター ---
    ws = wb['キャラクターアイディア']
    c = {h: col_of(ws, h) for h in ['ID', 'カード名', '作品名', 'レベル', 'サブ区分', '攻撃力',
                                    '没で得るIP', 'イチオシ', 'パッシブ能力', '技1名', '技1コスト',
                                    '技1効果', 'アイコン絵文字', 'イラストメモ']}
    for r in range(2, ws.max_row + 1):
        cid = nz(ws.cell(row=r, column=c['ID']).value)
        name = nz(ws.cell(row=r, column=c['カード名']).value)
        if not cid or not name:
            continue
        work = nz(ws.cell(row=r, column=c['作品名']).value)
        skill1 = nz(ws.cell(row=r, column=c['技1効果']).value)
        card = {
            'id': new_id(), 'cardCode': cid, 'name': name, 'type': 'キャラクター', 'work': work,
            'level': numstr(ws.cell(row=r, column=c['レベル']).value),
            'subrole': nz(ws.cell(row=r, column=c['サブ区分']).value),
            'atk': numstr(ws.cell(row=r, column=c['攻撃力']).value),
            'ip': numstr(ws.cell(row=r, column=c['没で得るIP']).value),
            'shinsaku': bool(nz(ws.cell(row=r, column=c['イチオシ']).value)),
            'passive': nz(ws.cell(row=r, column=c['パッシブ能力']).value),
            'skill1': skill1, 'skill1Effect': skill1,
            'skill1Name': nz(ws.cell(row=r, column=c['技1名']).value),
            'skill1Cost': numstr(ws.cell(row=r, column=c['技1コスト']).value),
            'icon': icon_for(name, ws.cell(row=r, column=c['アイコン絵文字']).value),
            'art': nz(ws.cell(row=r, column=c['イラストメモ']).value),
        }
        cards.append(card)
        if work:
            chars_by_work.setdefault(work, []).append(card)

    # --- 主人公 ---
    ws = wb['主人公']
    c = {h: col_of(ws, h) for h in ['ID', 'カード名', '作品名', '両面共通効果', '裏面効果名',
                                    'コスト', '裏面効果', 'アイコン絵文字', 'イラストメモ']}
    for r in range(2, ws.max_row + 1):
        cid = nz(ws.cell(row=r, column=c['ID']).value)
        name = nz(ws.cell(row=r, column=c['カード名']).value)
        if not cid or not name:
            continue
        work = nz(ws.cell(row=r, column=c['作品名']).value)
        card = {
            'id': new_id(), 'cardCode': cid, 'name': name, 'type': '主人公', 'work': work, 'level': '1',
            'frontEffect': nz(ws.cell(row=r, column=c['両面共通効果']).value),
            'backEffectName': nz(ws.cell(row=r, column=c['裏面効果名']).value),
            'backCost': numstr(ws.cell(row=r, column=c['コスト']).value),
            'backEffect': nz(ws.cell(row=r, column=c['裏面効果']).value),
            'icon': icon_for(name, ws.cell(row=r, column=c['アイコン絵文字']).value),
            'art': nz(ws.cell(row=r, column=c['イラストメモ']).value),
        }
        cards.append(card)
        if work:
            hero_by_work[work] = card['id']

    # --- ページカード / 急展開カード ---
    for sheet, typ in [('ページカード', 'ページカード'), ('急展開カード', '急展開カード')]:
        ws = wb[sheet]
        ci, cn, ce = col_of(ws, 'ID'), col_of(ws, 'カード名'), col_of(ws, '効果')
        for r in range(2, ws.max_row + 1):
            cid = nz(ws.cell(row=r, column=ci).value)
            name = nz(ws.cell(row=r, column=cn).value)
            if not cid or not name:
                continue
            card = {'id': new_id(), 'cardCode': cid, 'name': name, 'type': typ,
                    'effect1': nz(ws.cell(row=r, column=ce).value), 'icon': '', 'art': ''}
            cards.append(card)
            if typ == '急展開カード' and cid in ('KT-002', 'KT-003', 'KT-005'):
                kt_id[cid] = card['id']

    # --- タイプ(index.html BUILTIN_TYPES 相当)---
    types = [
        {'name': '主人公', 'color': '#a02d1f', 'builtin': True, 'kind': 'leveled', 'meaning': 'デッキアイコンカード (1 枚必須、デッキを識別)'},
        {'name': 'キャラクター', 'color': '#2c5fa0', 'builtin': True, 'kind': 'character', 'meaning': '場で戦う仲間'},
        {'name': '展開アイディア', 'color': '#3d8a3a', 'builtin': True, 'kind': 'effect', 'meaning': 'アイデア・物語の発展を促すカード'},
        {'name': 'テコ入れ', 'color': '#c8761f', 'builtin': True, 'kind': 'effect', 'meaning': '伏せて相手ターンに発動する介入カード'},
        {'name': 'どんでん返し', 'color': '#6e3a8b', 'builtin': True, 'kind': 'effect', 'meaning': '原稿6枚以上で使える大逆転カード'},
        {'name': 'ページカード', 'color': '#2f7d8a', 'builtin': True, 'kind': 'effect', 'meaning': 'v7: 毎ターンの共通ルール(ゲームの時計)'},
        {'name': '急展開カード', 'color': '#b03a6e', 'builtin': True, 'kind': 'effect', 'meaning': 'v7: 伏せて相手ターンにも撃てる介入カード'},
    ]

    # --- スターターデッキ(実カード参照・V7G- は作らない)---
    decks = []
    for work in ['ブラック・ジャック', '鉄腕アトム', '火の鳥', 'リボンの騎士']:
        entries = []
        if work in hero_by_work:
            entries.append({'cardId': hero_by_work[work], 'count': 1})
        namecount, added = {}, 0
        for card in chars_by_work.get(work, []):
            nm = card['name']
            if namecount.get(nm, 0) >= 6 or added >= 24:
                continue
            entries.append({'cardId': card['id'], 'count': 1})
            namecount[nm] = namecount.get(nm, 0) + 1
            added += 1
        for code in ['KT-003', 'KT-002', 'KT-005']:
            if code in kt_id:
                entries.append({'cardId': kt_id[code], 'count': 2})
        total = sum(e['count'] for e in entries)
        decks.append({'id': 'dseed_' + work, 'name': 'スターター: ' + work,
                      'description': '同梱スターターデッキ(編集可)', 'cards': entries, 'dondengCards': []})
        print('  デッキ:', work, total, '枚')

    seed = {'version': ver, 'generatedFrom': base, 'cards': cards, 'types': types,
            'decks': decks, 'cardIdCounter': seq[0] + 1}
    out = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', 'data', 'seed.js')
    out = os.path.normpath(out)
    with open(out, 'w', encoding='utf-8') as f:
        f.write('/* 自動生成: tools/build_seed.py(編集しないこと。xlsx を編集して再生成)*/\n')
        f.write('window.CARD_FORGE_SEED = ')
        f.write(json.dumps(seed, ensure_ascii=False, separators=(',', ':')))
        f.write(';\n')
    print('カード %d 枚 / タイプ %d / デッキ %d を書き出し → %s' % (len(cards), len(types), len(decks), out))


if __name__ == '__main__':
    main()
