# -*- coding: utf-8 -*-
"""手塚カードゲーム v7 — スターターデッキの「型番リスト」を生成する。

xlsx の「スタートデッキ」シートに、人が読める型番リスト(例: BJ-C001 2)を書き込む。
ユーザーはそれをコピーし、index.html の DECKS タブ「📋 型番リストから作成」に
貼り付けてデッキを作る。※カードは先に「📊 Excel取込」で取り込んでおくこと。

対象 xlsx: data/ 内の最新版 手塚カードゲーム_v7_vX.Y.xlsx
実行: py -3 tools/build_starter_decks.py
"""
import openpyxl, os, io, sys
from _common import latest_xlsx, backup

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

# ===== ブラック・ジャック スターターデッキ(主人公1 + メイン30 = 31枚)=====
DECK_NAME = 'ブラック・ジャック スターター「執刀」'
SECTIONS = [
    ('主人公',           [('BJ-H001', 1)]),
    ('ブラック・ジャック', [('BJ-C001', 2), ('BJ-C002', 1), ('BJ-C005', 2), ('BJ-C006', 1)]),
    ('ピノコ',           [('BJ-C009', 2), ('BJ-C010', 1), ('BJ-C013', 2), ('BJ-C014', 1)]),
    ('本間丈太郎',        [('BJ-C017', 2), ('BJ-C018', 1), ('BJ-C021', 2), ('BJ-C022', 1)]),
    ('ドクター・キリコ',  [('BJ-C025', 2), ('BJ-C026', 1), ('BJ-C029', 2), ('BJ-C030', 1)]),
    ('急展開カード',      [('KT-002', 2), ('KT-001', 1), ('KT-003', 1), ('KT-005', 1), ('KT-007', 1)]),
]
SHEET = 'スタートデッキ'


def collect_ids(ws):
    """シートの ID 列の値をすべて集める。"""
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return set()
    headers = [str(h).strip() if h is not None else '' for h in rows[0]]
    try:
        idc = headers.index('ID')
    except ValueError:
        idc = 0
    return {str(r[idc]).strip() for r in rows[1:]
            if idc < len(r) and r[idc] is not None and str(r[idc]).strip()}


def main():
    xlsx = latest_xlsx()
    print('対象:', xlsx)
    print('バックアップ:', backup(xlsx))

    wb = openpyxl.load_workbook(xlsx)

    # --- 検証: レシピの型番がすべて xlsx に存在するか ---
    known = set()
    for s in ('キャラクターアイディア', '主人公', '急展開カード'):
        known |= collect_ids(wb[s])
    total = 0
    missing = []
    for _label, entries in SECTIONS:
        for code, cnt in entries:
            total += cnt
            if code not in known:
                missing.append(code)
    if missing:
        print('ERROR: xlsx に存在しない型番:', ', '.join(missing))
        sys.exit(1)
    assert total == 31, '合計枚数が31でない: %d' % total
    print('検証OK: 全型番が xlsx に存在 / デッキ合計 %d 枚' % total)

    # --- 「スタートデッキ」シートの行を構築 ---
    lines = [
        '# ■ v7 スターターデッキ「ブラック・ジャック」',
        '# ① index.html で先に「📊 Excel取込」でカードを取り込む',
        '# ② このシートのセルをクリック → Ctrl+A → Ctrl+C で全コピー',
        '# ③ index.html → DECKS タブ →「📋 型番リストから作成」→ 貼り付け →「作成」',
        'デッキ名: ' + DECK_NAME,
    ]
    for label, entries in SECTIONS:
        lines.append('# --- %s ---' % label)
        for code, cnt in entries:
            lines.append('%s %d' % (code, cnt))

    # --- シート生成(既存なら作り直し)---
    if SHEET in wb.sheetnames:
        del wb[SHEET]
    ws = wb.create_sheet(SHEET)
    for i, ln in enumerate(lines, start=1):
        ws.cell(row=i, column=1).value = ln
    ws.column_dimensions['A'].width = 52

    wb.save(xlsx)
    print('「%s」シートを生成しました(%d行)' % (SHEET, len(lines)))
    print('保存完了:', xlsx)


if __name__ == '__main__':
    main()
