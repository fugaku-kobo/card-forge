# -*- coding: utf-8 -*-
"""手塚カードゲーム v7 — 攻撃力3桁化(×100)+ 貫通カード追加 マイグレーション。

data/ の最新版 xlsx(v1.1)を v1.2 にコピーし、次を行う(すべて冪等):
  1. キャラクターアイディアの「攻撃力」列(数値)を ×100。
  2. 全シートの効果テキスト内の攻撃修正(符号付き [+\\-−]N と「N以下」)を ×100。
     - ハイフン付きID(PG-001/KT-003)・日付(2026-05-22)は直前がASCII英数字のため除外。
  3. 詳細ルールの攻撃力レンジ式(Lv1:2〜4 / Lv2:4〜6)を ×100。
  4. 全キャラクター名に「貫通だけを持つ」カード(攻撃力500・パッシブ=貫通)を1枚ずつ追記。
  5. 詳細ルールに貫通キーワードの補足を追記。
冪等: ×100は「既に3桁(>=100)ならスキップ」、貫通行は既存なら追記しない。

実行順: 本スクリプト → tools/sync_xlsx_v7.py(文面を最新化)
実行: py -3 tools/migrate_x100_pierce.py
"""
import openpyxl, os, io, sys, re, shutil
from _common import latest_xlsx, backup, DATA_DIR

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

PIERCE_ATK = 300          # 貫通カードの攻撃力(扱いやすい中量級)
PIERCE_KEYWORD = '貫通'

# 符号付きの攻撃修正。直前がASCII英数字(ID/日付)の場合は対象外。
RE_SIGNED = re.compile(r'(?<![A-Za-z0-9])([+\-−])\s?(\d+)')
RE_IKA = re.compile(r'(\d+)(以下)')

changes = []


def x100_signed(s):
    """文字列内の符号付き攻撃修正と「N以下」を×100(<100のみ)。"""
    def rep_sign(m):
        n = int(m.group(2))
        return m.group(1) + (str(n * 100) if n < 100 else m.group(2))

    def rep_ika(m):
        n = int(m.group(1))
        return (str(n * 100) if n < 100 else m.group(1)) + m.group(2)

    return RE_IKA.sub(rep_ika, RE_SIGNED.sub(rep_sign, s))


def col_of(ws, header):
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=c).value
        if v is not None and str(v).strip() == header:
            return c
    return None


def last_data_row(ws, idc):
    last = 1
    for r in range(2, ws.max_row + 1):
        v = ws.cell(row=r, column=idc).value
        if v is not None and str(v).strip() != '':
            last = r
    return last


def has_pierce(path):
    """xlsx が既に移行済み(貫通カードを含む)かを判定。"""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb['キャラクターアイディア']
    passc = col_of(ws, 'パッシブ能力')
    if not passc:
        wb.close(); return False
    for row in ws.iter_rows(min_row=2, values_only=True):
        if passc - 1 < len(row) and str(row[passc - 1] or '') == PIERCE_KEYWORD:
            wb.close(); return True
    wb.close(); return False


def target_path():
    """移行対象パスを返す。未移行なら版を1つ上げてコピー、移行済みなら同版。"""
    cur = latest_xlsx()
    if has_pierce(cur):
        print('既に移行済み(貫通カードあり)→ 同版を対象:', os.path.basename(cur))
        return cur
    m = re.search(r'_v(\d+)\.(\d+)\.xlsx$', os.path.basename(cur))
    major, minor = int(m.group(1)), int(m.group(2))
    nxt = os.path.join(DATA_DIR, '手塚カードゲーム_v7_v%d.%d.xlsx' % (major, minor + 1))
    shutil.copy2(cur, nxt)
    print('版上げ: %s → %s' % (os.path.basename(cur), os.path.basename(nxt)))
    return nxt


def migrate_atk_column(ws):
    idc = col_of(ws, 'ID')
    atkc = col_of(ws, '攻撃力')
    if not atkc:
        return
    for r in range(2, ws.max_row + 1):
        cid = ws.cell(row=r, column=idc).value
        if cid is None or str(cid).strip() == '':
            continue
        cell = ws.cell(row=r, column=atkc)
        v = cell.value
        if isinstance(v, int) and 0 < v < 100:
            cell.value = v * 100
            changes.append('攻撃力 %s: %d→%d' % (cid, v, v * 100))


def migrate_text(ws):
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(row=r, column=c)
            v = cell.value
            if not isinstance(v, str):
                continue
            nv = x100_signed(v)
            # 攻撃力レンジ式(詳細ルール)
            nv = nv.replace('Lv1：2〜4 / Lv2：4〜6', 'Lv1：200〜400 / Lv2：400〜600')
            if nv != v:
                cell.value = nv
                changes.append('%s R%dC%d 文面×100' % (ws.title, r, c))


def add_pierce_cards(ws):
    idc = col_of(ws, 'ID')
    namec = col_of(ws, 'カード名')
    workc = col_of(ws, '作品名')
    lvc = col_of(ws, 'レベル')
    subc = col_of(ws, 'サブ区分')
    atkc = col_of(ws, '攻撃力')
    passc = col_of(ws, 'パッシブ能力')

    # 既に貫通カードがあれば追記しない(冪等)
    for r in range(2, ws.max_row + 1):
        if str(ws.cell(row=r, column=passc).value or '') == PIERCE_KEYWORD:
            print('貫通カードは既に存在 → 追記スキップ')
            return

    # Lv2(仕上げ)テンプレートのレベル/サブ区分表記を取得
    lv2_val, sub_val = '２', '仕上げ'
    # 各名前の代表情報(prefix, work)を収集(出現順を保持)
    order = []
    info = {}
    for r in range(2, ws.max_row + 1):
        cid = ws.cell(row=r, column=idc).value
        if cid is None or str(cid).strip() == '':
            continue
        cid = str(cid).strip()
        nm = str(ws.cell(row=r, column=namec).value or '').strip()
        wk = ws.cell(row=r, column=workc).value
        if not nm:
            continue
        if str(ws.cell(row=r, column=subc).value or '') == '仕上げ':
            lv2_val = ws.cell(row=r, column=lvc).value or lv2_val
        if nm not in info:
            prefix = cid.split('-')[0] if '-' in cid else 'GN'
            info[nm] = {'prefix': prefix, 'work': wk}
            order.append(nm)

    # prefixごとに -P## を採番
    pcount = {}
    start = last_data_row(ws, idc) + 1
    for i, nm in enumerate(order):
        pre = info[nm]['prefix']
        pcount[pre] = pcount.get(pre, 0) + 1
        new_id = '%s-P%02d' % (pre, pcount[pre])
        row = start + i
        ws.cell(row=row, column=idc).value = new_id
        ws.cell(row=row, column=namec).value = nm
        ws.cell(row=row, column=workc).value = info[nm]['work']
        ws.cell(row=row, column=lvc).value = lv2_val
        ws.cell(row=row, column=subc).value = sub_val
        ws.cell(row=row, column=atkc).value = PIERCE_ATK
        ws.cell(row=row, column=passc).value = PIERCE_KEYWORD
        changes.append('貫通カード追加 %s(%s) 攻撃力%d' % (new_id, nm, PIERCE_ATK))
        print('  + %s  %s  [%s]  ATK%d' % (new_id, nm, info[nm]['work'], PIERCE_ATK))


def ensure_pierce_rule(ws):
    """詳細ルールに貫通キーワードの補足を追記し、共通裁定行に貫通例外を注記(冪等)。"""
    # 既存の「盾持ち・守り手は代替判定でも有効」行に貫通例外を注記
    for r in range(1, ws.max_row + 1):
        s = str(ws.cell(row=r, column=1).value or '')
        if '盾持ち・守り手は代替判定でも有効' in s and '貫通' not in s:
            ws.cell(row=r, column=1).value = s.rstrip('。') + '。ただし貫通を持つアタッカーには無効。'
            changes.append('詳細ルール: 共通裁定行に貫通例外を注記')

    existing = [str(ws.cell(row=r, column=1).value or '')
                for r in range(1, ws.max_row + 1)]
    if any('「貫通」' in x for x in existing):
        return
    lines = [
        '・キーワード「貫通」(2026-06-07追加):攻撃時、相手の盾持ち・守り手の'
        '「完全に止める」防御を無視し、通常の攻撃力比べに持ち込む(勝てば突破して得点)。'
        '代替判定(PG-001〜005)下でも盾無視は有効。',
        '・攻撃力は3桁表記(2026-06-07に全数値×100)。',
    ]
    base = ws.max_row
    for i, line in enumerate(lines, start=1):
        ws.cell(row=base + i, column=1).value = line
    changes.append('詳細ルール: 貫通キーワードの補足を追記')


def main():
    xlsx = target_path()
    print('対象:', os.path.basename(xlsx))
    print('バックアップ:', backup(xlsx))

    wb = openpyxl.load_workbook(xlsx)
    ws_char = wb['キャラクターアイディア']
    migrate_atk_column(ws_char)
    for sn in wb.sheetnames:
        migrate_text(wb[sn])
    add_pierce_cards(ws_char)
    ensure_pierce_rule(wb['詳細ルール'])

    wb.save(xlsx)
    print('--- 変更 %d 件(先頭20件)---' % len(changes))
    for c in changes[:20]:
        print('  ', c)
    if len(changes) > 20:
        print('   ... 他 %d 件' % (len(changes) - 20))
    print('保存完了:', os.path.basename(xlsx))
    print('次に実行: py -3 tools/sync_xlsx_v7.py')


if __name__ == '__main__':
    main()
